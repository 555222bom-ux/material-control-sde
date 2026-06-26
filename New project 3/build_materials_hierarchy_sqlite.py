import datetime as dt
import os
import sqlite3
from pathlib import Path

import openpyxl

BASE_DIR = Path(r"C:\Users\bom\Documents\New project 3")
INPUT_XLSX = BASE_DIR / "outputs" / "db_split" / "ฐานข้อมูล_แยกตามหมวด.xlsx"
OUTPUT_DB = BASE_DIR / "outputs" / "db_split" / "materials_hierarchy.sqlite3"
SKIP_SHEETS = {"สรุป", "ตรวจสอบ"}

GROUP_MAP = {
    "A505": ("Asset", "Office_Equipment (อุปกรณ์สำนักงาน)"),
    "A510": ("Asset", "Safety_Site_Equipment (อุปกรณ์ความปลอดภัยและหน้างาน)"),
    "A515": ("Asset", "Safety_Site_Equipment (อุปกรณ์ความปลอดภัยและหน้างาน)"),
    "A520": ("Asset", "Safety_Site_Equipment (อุปกรณ์ความปลอดภัยและหน้างาน)"),
    "A600": ("Asset", "Tools_Testing_Equipment (เครื่องมือและเครื่องทดสอบ)"),
    "A601": ("Asset", "Tools_Testing_Equipment (เครื่องมือและเครื่องทดสอบ)"),
    "A610": ("Asset", "Tools_Testing_Equipment (เครื่องมือและเครื่องทดสอบ)"),
    "P001": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P002": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P003": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P004": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P005": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P006": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P007": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P009": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P010": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P011": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P013": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P014": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P015": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P020": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P021": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P022": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P025": ("Product", "Solar_Power_Materials (วัสดุระบบโซลาร์)"),
    "P026": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P027": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P029": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P031": ("Product", "Electrical_Equipment_System (อุปกรณ์และระบบไฟฟ้า)"),
    "P032": ("Product", "Cable_Wire_Accessories (สายไฟและอุปกรณ์ประกอบ)"),
    "P033": ("Product", "Cable_Wire_Accessories (สายไฟและอุปกรณ์ประกอบ)"),
    "P034": ("Product", "Conduit_Raceway_Support (ท่อร้อยสาย ราง และโครงรับ)"),
    "P035": ("Product", "Conduit_Raceway_Support (ท่อร้อยสาย ราง และโครงรับ)"),
    "P036": ("Product", "Conduit_Raceway_Support (ท่อร้อยสาย ราง และโครงรับ)"),
    "P037": ("Product", "Conduit_Raceway_Support (ท่อร้อยสาย ราง และโครงรับ)"),
    "P038": ("Product", "Conduit_Raceway_Support (ท่อร้อยสาย ราง และโครงรับ)"),
    "P039": ("Product", "Conduit_Raceway_Support (ท่อร้อยสาย ราง และโครงรับ)"),
    "P040": ("Product", "Fastener_Hardware (น็อต สกรู และฮาร์ดแวร์)"),
    "P041": ("Product", "Paint_Construction Chemical (สีและวัสดุเคมี)"),
    "P042": ("Product", "Paint_Construction Chemical (สีและวัสดุเคมี)"),
    "P043": ("Product", "Pipe_Plumbing_Materials (ท่อและงานประปา)"),
    "P100": ("Product", "Mechanical_Pneumatic_Instrument (เครื่องกล ลม และเครื่องมือวัด)"),
    "P105": ("Product", "Mechanical_Pneumatic_Instrument (เครื่องกล ลม และเครื่องมือวัด)"),
    "P106": ("Product", "Mechanical_Pneumatic_Instrument (เครื่องกล ลม และเครื่องมือวัด)"),
    "P110": ("Product", "Pipe_Plumbing_Materials (ท่อและงานประปา)"),
    "P210": ("Product", "HVAC_Ventilation (ระบบระบายอากาศ)"),
    "P215": ("Product", "HVAC_Ventilation (ระบบระบายอากาศ)"),
    "P220": ("Product", "Pipe_Plumbing_Materials (ท่อและงานประปา)"),
    "P280": ("Product", "Pump_Water_System (ปั๊มและระบบน้ำ)"),
    "P500": ("Product", "Consumable_General_Supply (วัสดุสิ้นเปลืองและของใช้ทั่วไป)"),
    "P501": ("Product", "Uniform_PPE (เครื่องแต่งกายและ PPE)"),
    "P502": ("Product", "Consumable_General_Supply (วัสดุสิ้นเปลืองและของใช้ทั่วไป)"),
    "P503": ("Product", "Uniform_PPE (เครื่องแต่งกายและ PPE)"),
    "P510": ("Product", "Uniform_PPE (เครื่องแต่งกายและ PPE)"),
    "P620": ("Product", "Tools_Consumable_Parts (อะไหล่และเครื่องมือสิ้นเปลือง)"),
    "P630": ("Product", "Insulation_Fiberglass_Materials (วัสดุฉนวนและไฟเบอร์กลาส)"),
    "PS01": ("Service", "Installation_Service (งานติดตั้ง)"),
    "PS04": ("Service", "Installation_Service (งานติดตั้ง)"),
    "PS05": ("Service", "Scaffolding_Service (งานนั่งร้าน)"),
    "PS07": ("Service", "Calibration_Service (งานสอบเทียบเครื่องมือ)"),
    "PS10": ("Service", "General_Service (ค่าบริการทั่วไป)"),
    "PS20": ("Service", "Other_Service (บริการอื่น ๆ)"),
}


def normalize(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def category_from_code(material_code):
    code = normalize(material_code)
    if not code:
        return "ไม่ระบุ"
    return code.upper()[:4]


def best_sub_group_name(category, rows):
    for row in rows:
        code = normalize(row["material_code"]) or ""
        name = normalize(row["material_name"])
        tail = code[len(category) :]
        if name and tail and set(tail[:-1] or "0") <= {"0"}:
            return name
    return rows[0]["material_name"] or category


def read_grouped_rows():
    workbook = openpyxl.load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    grouped = {}
    for sheet_name in workbook.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
            if not row or not any(cell is not None for cell in row):
                continue
            material_code = normalize(row[2])
            if not material_code:
                continue
            category = normalize(row[0]) or category_from_code(material_code)
            grouped.setdefault(category, []).append(
                {
                    "category": category,
                    "source_no": int(row[1]) if isinstance(row[1], (int, float)) and row[1] == int(row[1]) else row[1],
                    "material_code": material_code,
                    "material_name": normalize(row[3]),
                    "spec_size": normalize(row[4]),
                    "brand": normalize(row[5]),
                    "unit": normalize(row[6]),
                }
            )
    return grouped


def rebuild_database(grouped):
    if OUTPUT_DB.exists():
        OUTPUT_DB.unlink()

    conn = sqlite3.connect(OUTPUT_DB)
    conn.execute("PRAGMA foreign_keys=ON")
    conn.executescript(
        """
        CREATE TABLE material_types (
            type_name TEXT PRIMARY KEY
        );

        CREATE TABLE material_groups (
            group_id INTEGER PRIMARY KEY AUTOINCREMENT,
            type_name TEXT NOT NULL,
            group_name TEXT NOT NULL,
            UNIQUE(type_name, group_name),
            FOREIGN KEY(type_name) REFERENCES material_types(type_name)
        );

        CREATE TABLE material_sub_groups (
            category_code TEXT PRIMARY KEY,
            group_id INTEGER NOT NULL,
            sub_group_name TEXT NOT NULL,
            item_count INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY(group_id) REFERENCES material_groups(group_id)
        );

        CREATE TABLE materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category_code TEXT NOT NULL,
            source_no INTEGER,
            material_code TEXT NOT NULL UNIQUE,
            material_name TEXT,
            spec_size TEXT,
            brand TEXT,
            unit TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(category_code) REFERENCES material_sub_groups(category_code)
        );

        CREATE INDEX idx_materials_category_code ON materials(category_code);
        CREATE INDEX idx_materials_code ON materials(material_code);
        CREATE INDEX idx_materials_name ON materials(material_name);
        CREATE INDEX idx_materials_brand ON materials(brand);
        CREATE INDEX idx_sub_groups_group_id ON material_sub_groups(group_id);

        CREATE VIEW material_catalog AS
        SELECT
            mt.type_name AS type,
            mg.group_name AS group_name,
            msg.category_code,
            msg.sub_group_name,
            m.material_code,
            m.material_name,
            m.spec_size,
            m.brand,
            m.unit
        FROM materials m
        JOIN material_sub_groups msg ON msg.category_code = m.category_code
        JOIN material_groups mg ON mg.group_id = msg.group_id
        JOIN material_types mt ON mt.type_name = mg.type_name;

        CREATE VIEW brand_catalog AS
        SELECT
            type,
            group_name,
            sub_group_name,
            brand,
            COUNT(*) AS item_count
        FROM material_catalog
        WHERE brand IS NOT NULL
        GROUP BY type, group_name, sub_group_name, brand;

        CREATE TABLE import_metadata (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        """
    )

    group_id_cache = {}
    rows_to_insert = []
    with conn:
        for category in sorted(grouped):
            type_name, group_name = GROUP_MAP.get(category, ("Product", "Unmapped (รอจัดหมวด)"))
            sub_group_name = best_sub_group_name(category, grouped[category])
            conn.execute("INSERT OR IGNORE INTO material_types(type_name) VALUES (?)", (type_name,))
            conn.execute(
                "INSERT OR IGNORE INTO material_groups(type_name, group_name) VALUES (?, ?)",
                (type_name, group_name),
            )
            group_id = group_id_cache.get((type_name, group_name))
            if group_id is None:
                group_id = conn.execute(
                    "SELECT group_id FROM material_groups WHERE type_name = ? AND group_name = ?",
                    (type_name, group_name),
                ).fetchone()[0]
                group_id_cache[(type_name, group_name)] = group_id
            conn.execute(
                """
                INSERT INTO material_sub_groups(category_code, group_id, sub_group_name, item_count)
                VALUES (?, ?, ?, ?)
                """,
                (category, group_id, sub_group_name, len(grouped[category])),
            )
            rows_to_insert.extend(grouped[category])

        conn.executemany(
            """
            INSERT INTO materials (
                category_code, source_no, material_code, material_name, spec_size, brand, unit
            ) VALUES (
                :category, :source_no, :material_code, :material_name, :spec_size, :brand, :unit
            )
            """,
            rows_to_insert,
        )
        conn.executemany(
            "INSERT INTO import_metadata(key, value) VALUES (?, ?)",
            [
                ("source_file", str(INPUT_XLSX)),
                ("created_at", dt.datetime.now().isoformat(timespec="seconds")),
                ("material_rows", str(len(rows_to_insert))),
                ("sub_group_rows", str(len(grouped))),
                ("group_rows", str(len(group_id_cache))),
            ],
        )
    conn.execute("PRAGMA optimize")
    return conn


def verify(conn):
    checks = {
        "materials": conn.execute("SELECT COUNT(*) FROM materials").fetchone()[0],
        "unique_codes": conn.execute("SELECT COUNT(DISTINCT material_code) FROM materials").fetchone()[0],
        "types": conn.execute("SELECT COUNT(*) FROM material_types").fetchone()[0],
        "groups": conn.execute("SELECT COUNT(*) FROM material_groups").fetchone()[0],
        "sub_groups": conn.execute("SELECT COUNT(*) FROM material_sub_groups").fetchone()[0],
        "duplicate_codes": conn.execute(
            """
            SELECT COUNT(*)
            FROM (
                SELECT material_code
                FROM materials
                GROUP BY material_code
                HAVING COUNT(*) > 1
            )
            """
        ).fetchone()[0],
        "paint_rows": conn.execute(
            """
            SELECT COUNT(*)
            FROM material_catalog
            WHERE type = 'Product'
              AND group_name = 'Paint_Construction Chemical (สีและวัสดุเคมี)'
              AND sub_group_name LIKE 'Paint%'
            """
        ).fetchone()[0],
    }
    sample_paint = conn.execute(
        """
        SELECT type, group_name, sub_group_name, material_code, brand
        FROM material_catalog
        WHERE group_name = 'Paint_Construction Chemical (สีและวัสดุเคมี)'
        ORDER BY category_code, material_code
        LIMIT 5
        """
    ).fetchall()
    return checks, sample_paint


def main():
    grouped = read_grouped_rows()
    conn = rebuild_database(grouped)
    try:
        checks, sample_paint = verify(conn)
        print(f"output={OUTPUT_DB}")
        print(f"size_kb={os.path.getsize(OUTPUT_DB) / 1024:.1f}")
        print(f"checks={checks}")
        print(f"sample_paint={sample_paint}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
