import datetime as dt
import os
import sqlite3
from pathlib import Path

import openpyxl

BASE_DIR = Path(r"C:\Users\bom\Documents\New project 3")
INPUT_XLSX = BASE_DIR / "outputs" / "db_split" / "ฐานข้อมูล_แยกตามหมวด.xlsx"
OUTPUT_DB = BASE_DIR / "outputs" / "db_split" / "materials.sqlite3"

SKIP_SHEETS = {"สรุป", "ตรวจสอบ"}


def normalize(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def expected_category(material_code):
    code = normalize(material_code)
    if not code:
        return "ไม่ระบุ"
    return code.upper()[:4]


def read_rows():
    workbook = openpyxl.load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(cell is not None for cell in row):
                continue
            category = normalize(row[0]) or expected_category(row[2])
            source_no = row[1]
            material_code = normalize(row[2])
            if not material_code:
                continue
            yield {
                "category": category,
                "source_no": int(source_no) if isinstance(source_no, (int, float)) and source_no == int(source_no) else source_no,
                "material_code": material_code,
                "material_name": normalize(row[3]),
                "spec_size": normalize(row[4]),
                "brand": normalize(row[5]),
                "unit": normalize(row[6]),
            }


def build_database(rows):
    if OUTPUT_DB.exists():
        OUTPUT_DB.unlink()

    conn = sqlite3.connect(OUTPUT_DB)
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA foreign_keys=ON")

        conn.executescript(
            """
            CREATE TABLE categories (
                category TEXT PRIMARY KEY,
                item_count INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE materials (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT NOT NULL,
                source_no INTEGER,
                material_code TEXT NOT NULL UNIQUE,
                material_name TEXT,
                spec_size TEXT,
                brand TEXT,
                unit TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (category) REFERENCES categories(category)
            );

            CREATE INDEX idx_materials_category ON materials(category);
            CREATE INDEX idx_materials_code ON materials(material_code);
            CREATE INDEX idx_materials_name ON materials(material_name);
            CREATE INDEX idx_materials_brand ON materials(brand);

            CREATE VIEW materials_light AS
            SELECT
                category,
                material_code,
                material_name,
                spec_size,
                brand,
                unit
            FROM materials;

            CREATE TABLE import_metadata (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            """
        )

        rows = list(rows)
        categories = sorted({row["category"] for row in rows})

        with conn:
            conn.executemany(
                "INSERT INTO categories(category) VALUES (?)",
                [(category,) for category in categories],
            )
            conn.executemany(
                """
                INSERT INTO materials (
                    category, source_no, material_code, material_name, spec_size, brand, unit
                ) VALUES (
                    :category, :source_no, :material_code, :material_name, :spec_size, :brand, :unit
                )
                """,
                rows,
            )
            conn.execute(
                """
                UPDATE categories
                SET item_count = (
                    SELECT COUNT(*)
                    FROM materials
                    WHERE materials.category = categories.category
                )
                """
            )
            conn.executemany(
                "INSERT INTO import_metadata(key, value) VALUES (?, ?)",
                [
                    ("source_file", str(INPUT_XLSX)),
                    ("created_at", dt.datetime.now().isoformat(timespec="seconds")),
                    ("material_rows", str(len(rows))),
                    ("category_rows", str(len(categories))),
                ],
            )

        conn.execute("PRAGMA optimize")
        return conn
    except Exception:
        conn.close()
        raise


def verify(conn):
    checks = {
        "materials": conn.execute("SELECT COUNT(*) FROM materials").fetchone()[0],
        "unique_codes": conn.execute("SELECT COUNT(DISTINCT material_code) FROM materials").fetchone()[0],
        "categories": conn.execute("SELECT COUNT(*) FROM categories").fetchone()[0],
        "duplicate_codes": conn.execute(
            """
            SELECT COUNT(*)
            FROM (
                SELECT material_code
                FROM materials
                GROUP BY material_code
                HAVING COUNT(*) > 1
            )
            """
        ).fetchone()[0],
        "wrong_category": conn.execute(
            """
            SELECT COUNT(*)
            FROM materials
            WHERE category != UPPER(SUBSTR(material_code, 1, 4))
            """
        ).fetchone()[0],
    }
    sample = conn.execute(
        """
        SELECT category, item_count
        FROM categories
        ORDER BY category
        LIMIT 10
        """
    ).fetchall()
    return checks, sample


def main():
    rows = list(read_rows())
    conn = build_database(rows)
    try:
        checks, sample = verify(conn)
        size_kb = os.path.getsize(OUTPUT_DB) / 1024
        print(f"output={OUTPUT_DB}")
        print(f"size_kb={size_kb:.1f}")
        print(f"checks={checks}")
        print(f"sample_categories={sample}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
