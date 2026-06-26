import datetime
import json
import pathlib
import re
import sqlite3
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = pathlib.Path(__file__).resolve().parent

SRC_DB = (
    BASE_DIR / "New project 3" / "outputs" / "db_split" / "materials_hierarchy.sqlite3"
)
OUT_JSON = BASE_DIR / "MaterialControlData_from_New_project_3.json"
OUT_XLSX = BASE_DIR / "MasterData_Import_from_New_project_3.xlsx"


def clean(value):
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) >= 2 and text[0] == text[-1] == '"':
        text = text[1:-1].strip()
    return text


def make_abbr(code, name):
    raw = clean(code) or clean(name)[:2].upper() or "GN"
    raw = re.sub(r"[^A-Za-z0-9_-]", "", raw).upper()
    return raw or "GN"


def make_category_name(code, name):
    clean_name = clean(name) or clean(code) or "General"
    return clean_name


def clean_unit(value):
    unit = clean(value)
    if unit == "EAB2C2147:F2537":
        return "EA"
    return unit or "Pcs"


def sheet_name_for_category(category):
    name = clean(category).split(" - ", 1)[0] or "General"
    name = re.sub(r"[\\/*?:\[\]]", "_", name)
    return name[:31] or "General"


def write_master_import_workbook(master_data):
    wb = Workbook()
    wb.remove(wb.active)

    groups = OrderedDict()
    for item in master_data:
        groups.setdefault(sheet_name_for_category(item["category"]), []).append(item)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for sheet_name, rows in groups.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(["Code", "Name", "Spec/Size", "Brand", "Unit", "BOQ"])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for item in rows:
            ws.append(
                [
                    item["code"],
                    item["name"],
                    item["spec"],
                    item["brand"],
                    item["unit"],
                    item["boq"],
                ]
            )

        for col, width in zip("ABCDEF", [18, 48, 42, 20, 12, 10]):
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"

    wb.save(OUT_XLSX)
    return len(groups)


def main():
    con = sqlite3.connect(SRC_DB)
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    rows = cur.execute(
        """
        select
            m.id,
            m.material_code,
            m.material_name,
            m.spec_size,
            m.brand,
            m.unit,
            m.category_code,
            coalesce(sg.sub_group_name, m.category_code) as category_name
        from materials m
        left join material_sub_groups sg on sg.category_code = m.category_code
        order by m.id
        """
    ).fetchall()

    cats = cur.execute(
        """
        select
            sg.category_code,
            coalesce(sg.sub_group_name, sg.category_code) as category_name
        from material_sub_groups sg
        order by sg.category_code
        """
    ).fetchall()

    units = cur.execute(
        "select distinct unit from materials where unit is not null and trim(unit) <> '' order by unit"
    ).fetchall()

    category_list = []
    seen_categories = set()
    for idx, row in enumerate(cats, 1):
        name = make_category_name(row["category_code"], row["category_name"])
        key = name.casefold()
        if key in seen_categories:
            continue
        seen_categories.add(key)
        category_list.append(
            {
                "id": f"CAT-{idx}",
                "name": name,
                "abbr": make_abbr(row["category_code"], name),
            }
        )

    unit_list = []
    seen_units = set()
    for idx, row in enumerate(units, 1):
        name = clean_unit(row["unit"])
        key = name.casefold()
        if key in seen_units:
            continue
        seen_units.add(key)
        unit_list.append({"id": f"UNIT-{idx}", "name": name})
    if not unit_list:
        unit_list = [{"id": "UNIT-1", "name": "Pcs"}]

    master_data = []
    for idx, row in enumerate(rows, 1):
        name = clean(row["material_name"])
        if not name:
            continue
        master_data.append(
            {
                "id": f"MAT-{idx}",
                "code": clean(row["material_code"]),
                "name": name,
                "brand": clean(row["brand"]),
                "spec": clean(row["spec_size"]) or "-",
                "unit": clean_unit(row["unit"]),
                "category": make_category_name(row["category_code"], row["category_name"]),
                "boq": 0,
                "ordered": [],
                "received": [],
                "disbursed": [],
            }
        )

    data = {
        "reqConfig": {"prefix": "REQ-24-", "nextRun": 1},
        "masterData": master_data,
        "categoryList": category_list or [{"id": "CAT-1", "name": "General", "abbr": "GN"}],
        "unitList": unit_list,
        "personnel": [],
        "requests": [],
        "receipts": [],
        "plans": [],
        "vendors": [],
        "companies": [{"id": "COMPANY-1", "name": "บริษัทหลัก (Main Contractor)"}],
        "source": {
            "from": str(SRC_DB),
            "createdAt": datetime.datetime.now().isoformat(timespec="seconds"),
            "materials": len(master_data),
        },
    }

    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    sheet_count = write_master_import_workbook(master_data)
    con.close()

    print(
        json.dumps(
            {
                "output": str(OUT_JSON),
                "xlsx": str(OUT_XLSX),
                "materials": len(master_data),
                "categories": len(data["categoryList"]),
                "units": len(unit_list),
                "excelSheets": sheet_count,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
